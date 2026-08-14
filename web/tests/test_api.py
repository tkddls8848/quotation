"""`POST /api/v1/convert` 와 보조 엔드포인트의 계약 (계획서 §6, §12.2)."""
from __future__ import annotations

import datetime as dt
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook

import api
import clock
import errors
import limits

TODAY = dt.date(2026, 7, 23)


def _convert(uploads, template_bytes, **kwargs) -> api.ApiResponse:
    params = dict(template_bytes=template_bytes, template_version="v-test",
                  deployment_version="d-test", request_id="req-1", today=TODAY)
    params.update(kwargs)
    return api.convert_response(uploads, **params)


def _upload(fixtures, name, *, filename=None, content_type="text/xml"):
    return api.Upload(filename=filename or name,
                      content=(fixtures / name).read_bytes(),
                      content_type=content_type)


# --- 성공 --------------------------------------------------------------------

def test_convert_returns_a_downloadable_xlsx(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml")], template_bytes)

    assert res.status == 200
    assert res.headers["Content-Type"] == api.XLSX_CONTENT_TYPE
    assert res.headers["Cache-Control"] == "no-store"
    assert res.headers["X-Content-Type-Options"] == "nosniff"
    assert res.headers["X-Request-Id"] == "req-1"
    assert res.headers["X-Template-Version"] == "v-test"
    assert res.headers["Content-Length"] == str(len(res.body))

    wb = load_workbook(BytesIO(res.body))
    assert wb.sheetnames == ["TOTAL", "SAMPLE-100 #1", "SAMPLE-200 #2", "template"]
    assert wb["TOTAL"]["C3"].value == TODAY.isoformat()


def test_convert_keeps_the_logo_drawing(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml")], template_bytes)
    with zipfile.ZipFile(BytesIO(res.body)) as z:
        assert [n for n in z.namelist() if n.startswith("xl/media/")]


def test_download_name_keeps_the_stem(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml",
                            filename="견적 (수정본)#2.xml")], template_bytes)
    disposition = res.headers["Content-Disposition"]
    assert disposition.startswith("attachment; ")
    # 한글 이름은 RFC 5987 로, 구형 클라이언트용 ASCII 대체 이름도 함께 준다
    assert "filename*=UTF-8''" in disposition
    assert "%23" in disposition and ".xlsx" in disposition
    assert 'filename="' in disposition


@pytest.mark.parametrize("raw,expected", [
    ("../../etc/passwd.xml", "passwd.xml"),
    (r"C:\\Users\\me\\견적.xml", "견적.xml"),
    ("", "quotation.xml"),
    ("보고서\r\n.xml", "보고서.xml"),
])
def test_upload_names_are_stripped_of_paths(raw, expected):
    assert api.safe_source_name(raw) == expected


def test_euckr_upload_is_accepted(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "euckr_quote.xml")], template_bytes)
    assert res.status == 200
    assert "서버 1" in load_workbook(BytesIO(res.body)).sheetnames


def test_success_is_logged_without_sensitive_fields(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml",
                            filename="고객사_견적.xml")], template_bytes)
    assert res.log["outcome"] == "ok"
    assert res.log["group_count"] == 2
    assert res.log["line_count"] == 3
    assert res.log["input_size_bucket"] == "<=64KiB"
    assert res.log["template_version"] == "v-test"
    dumped = repr(res.log)
    assert "고객사" not in dumped, "파일명이 로그에 남았다"
    assert "SAMPLE" not in dumped, "제품 정보가 로그에 남았다"


# --- 입력 오류 ----------------------------------------------------------------

def _error(res: api.ApiResponse) -> dict:
    return res.json()["error"]


def test_missing_file_is_a_400(template_bytes):
    res = _convert([], template_bytes)
    assert res.status == 400
    assert _error(res)["code"] == errors.INVALID_REQUEST
    assert _error(res)["request_id"] == "req-1"
    assert res.headers["Cache-Control"] == "no-store"


def test_more_than_one_file_is_rejected(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml"),
                    _upload(fixtures, "no_charge.xml")], template_bytes)
    assert res.status == 400
    assert _error(res)["code"] == errors.INVALID_REQUEST


def test_non_xml_extension_is_a_415(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml", filename="quote.txt")],
                   template_bytes)
    assert res.status == 415
    assert _error(res)["code"] == errors.UNSUPPORTED_MEDIA_TYPE


def test_html_content_type_is_a_415(fixtures, template_bytes):
    res = _convert([_upload(fixtures, "new_quote.xml", content_type="text/html")],
                   template_bytes)
    assert res.status == 415


def test_oversized_upload_is_a_413(template_bytes):
    big = api.Upload(filename="big.xml", content=b"<CFXML/>" +
                     b"x" * limits.MAX_UPLOAD_BYTES, content_type="text/xml")
    res = _convert([big], template_bytes)
    assert res.status == 413
    assert _error(res)["code"] == errors.FILE_TOO_LARGE


def test_empty_file_is_a_400(template_bytes):
    res = _convert([api.Upload(filename="a.xml", content=b"   ",
                               content_type="text/xml")], template_bytes)
    assert res.status == 400


def test_malformed_xml_is_a_422(template_bytes):
    res = _convert([api.Upload(filename="a.xml", content=b"<CFXML><CFData>",
                               content_type="text/xml")], template_bytes)
    assert res.status == 422
    assert _error(res)["code"] == errors.INVALID_QUOTATION_XML


def test_document_without_quotable_items_is_a_422(template_bytes):
    body = b"<CFXML><CFData/></CFXML>"
    res = _convert([api.Upload(filename="a.xml", content=body,
                               content_type="text/xml")], template_bytes)
    assert res.status == 422
    # 원본 프로그램과 같은 문구를 그대로 보여 준다
    assert "Item을 찾을 수 없습니다" in _error(res)["message"]


def test_too_many_line_items_is_rejected_before_parsing(template_bytes):
    body = (b"<CFXML><CFData>"
            + b"<ProductLineItem></ProductLineItem>" * (limits.MAX_LINE_ITEMS + 1)
            + b"</CFData></CFXML>")
    res = _convert([api.Upload(filename="a.xml", content=body,
                               content_type="text/xml")], template_bytes)
    assert res.status == 422
    assert "품목이 너무 많습니다" in _error(res)["message"]


def test_too_many_groups_is_rejected(fixtures, template_bytes, monkeypatch):
    monkeypatch.setattr(limits, "MAX_GROUPS", 1)
    res = _convert([_upload(fixtures, "new_quote.xml")], template_bytes)
    assert res.status == 422
    assert "장비군이 너무 많습니다" in _error(res)["message"]


def test_external_entities_do_not_leak(tmp_path, template_bytes):
    secret = tmp_path / "secret.txt"
    secret.write_text("TOPSECRET", encoding="utf-8")
    body = (
        '<?xml version="1.0"?>'
        f'<!DOCTYPE CFXML [<!ENTITY xxe SYSTEM "file:///{secret.as_posix()}">]>'
        "<CFXML><CFData><ProductLineItem>"
        "<ProductLineNumber>&xxe;</ProductLineNumber>"
        "<TransactionType>NEW</TransactionType>"
        "</ProductLineItem></CFData></CFXML>"
    ).encode("utf-8")
    res = _convert([api.Upload(filename="a.xml", content=body,
                               content_type="text/xml")], template_bytes)
    assert b"TOPSECRET" not in res.body


def test_error_bodies_hide_internals(template_bytes):
    res = _convert([api.Upload(filename="a.xml", content=b"<CFXML><CFData>",
                               content_type="text/xml")], template_bytes)
    assert set(res.json()) == {"error"}
    assert set(_error(res)) == {"code", "message", "request_id"}
    message = _error(res)["message"]
    for leak in ("Traceback", "quotation/core", "/root", "R2", "openpyxl"):
        assert leak not in message


# --- 템플릿 오류 --------------------------------------------------------------

@pytest.mark.parametrize("template", [b"", b"not a zip"])
def test_unusable_template_is_a_503(fixtures, template):
    res = _convert([_upload(fixtures, "new_quote.xml")], template)
    assert res.status == 503
    assert _error(res)["code"] == errors.TEMPLATE_UNAVAILABLE


def test_template_missing_required_sheets_is_a_503(fixtures, template_bytes):
    """필수 시트가 없는 템플릿으로는 견적서를 만들 수 없다."""
    import conversion_adapter
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "TOTAL"          # 숨김 template 시트가 없다
    buffer = BytesIO()
    wb.save(buffer)

    with pytest.raises(errors.ApiError) as caught:
        conversion_adapter.validate_template(buffer.getvalue())
    assert caught.value.code == errors.TEMPLATE_UNAVAILABLE

    res = _convert([_upload(fixtures, "new_quote.xml")], buffer.getvalue())
    assert res.status == 503


# --- 보조 엔드포인트 -----------------------------------------------------------

def test_config_reports_client_side_limits():
    payload = api.config_response("req-2").json()
    assert payload["max_upload_bytes"] == limits.MAX_UPLOAD_BYTES
    assert payload["allowed_suffixes"] == [".xml"]
    assert payload["max_file_count"] == 1
    assert payload["max_batch_files"] == limits.MAX_BATCH_FILES


def test_the_screen_carries_the_same_limits():
    """화면은 상한을 물어보지 않고 빌드 시점 사본을 쓴다 (무료 계정의 요청을
    아끼기 위해서다). 그 사본이 여기 값과 어긋나면 안 된다."""
    import re
    from pathlib import Path

    source = (Path(__file__).resolve().parents[2] / "web" / "frontend" / "src"
              / "api.ts").read_text(encoding="utf-8")
    block = re.search(r"APP_CONFIG: AppConfig = \{(.*?)\};", source, re.S)
    assert block, "api.ts 에서 APP_CONFIG 를 찾지 못했습니다"
    body = block.group(1)

    def value(name: str) -> str:
        found = re.search(rf"{name}:\s*([^,\n]+)", body)
        assert found, f"api.ts 에 {name} 이 없습니다"
        return found.group(1).strip()

    assert eval(value("max_upload_bytes")) == limits.MAX_UPLOAD_BYTES
    assert int(value("max_file_count")) == limits.MAX_FILE_COUNT
    assert int(value("max_batch_files")) == limits.MAX_BATCH_FILES
    assert "'.xml'" in value("allowed_suffixes")


def test_status_reports_versions_only():
    payload = api.status_response("req-3", deployment_version="1.2.3",
                                  template_version="v9").json()
    assert payload == {"deployment_version": "1.2.3", "template_version": "v9"}


def test_method_not_allowed_lists_the_allowed_method():
    res = api.method_not_allowed("req-4", "POST")
    assert res.status == 405 and res.headers["Allow"] == "POST"


# --- 동일 출처 ----------------------------------------------------------------

@pytest.mark.parametrize("origin,site", [
    (None, "same-origin"),
    ("https://quote.example.com", "same-origin"),
    (None, None),
])
def test_same_origin_requests_pass(origin, site):
    api.check_same_origin(request_url="https://quote.example.com/api/v1/convert",
                          origin=origin, sec_fetch_site=site)


@pytest.mark.parametrize("origin,site", [
    ("https://evil.example.com", "cross-site"),
    ("https://evil.example.com", None),
    (None, "cross-site"),
])
def test_cross_site_requests_are_rejected(origin, site):
    with pytest.raises(errors.ApiError) as caught:
        api.check_same_origin(
            request_url="https://quote.example.com/api/v1/convert",
            origin=origin, sec_fetch_site=site)
    assert caught.value.status == 400


# --- 견적 날짜 ----------------------------------------------------------------

@pytest.mark.parametrize("utc,expected", [
    ("2026-08-13T14:59:00+00:00", dt.date(2026, 8, 13)),
    ("2026-08-13T15:00:00+00:00", dt.date(2026, 8, 14)),  # 서울은 이미 다음 날
    ("2026-08-13T23:30:00+00:00", dt.date(2026, 8, 14)),
])
def test_quote_date_follows_seoul_not_utc(utc, expected):
    assert clock.seoul_today(dt.datetime.fromisoformat(utc)) == expected


def test_seoul_date_is_used_for_the_sheet(fixtures, template_bytes):
    """UTC 로는 13일이지만 서울은 14일인 시각."""
    today = clock.seoul_today(dt.datetime.fromisoformat("2026-08-13T16:00:00+00:00"))
    res = _convert([_upload(fixtures, "new_quote.xml")], template_bytes,
                   today=today)
    assert load_workbook(BytesIO(res.body))["TOTAL"]["C3"].value == "2026-08-14"


# --- 번들 내장 템플릿 ----------------------------------------------------------

def test_bundled_template_matches_the_repository_original(template_bytes):
    """번들 템플릿은 저장소 원본과 바이트가 같아야 한다.

    R2 를 쓰지 않고 배포 직전에 sync_core.py 가 만들어 넣는다. 이 검증이 없으면
    낡은 사본이 조용히 배포될 수 있다.
    """
    import template

    assert template.template_bytes() == template_bytes
    assert template.template_version().startswith("sha256-")


def test_bundled_template_passes_validation():
    import conversion_adapter
    import template

    conversion_adapter.validate_template(template.template_bytes())


def test_convert_with_the_bundled_template(fixtures):
    """운영과 같은 경로: 업로드 + 번들 템플릿 -> 견적서."""
    import template

    res = _convert([_upload(fixtures, "new_quote.xml")], template.template_bytes(),
                   template_version=template.template_version())
    assert res.status == 200
    assert res.headers["X-Template-Version"].startswith("sha256-")
    assert load_workbook(BytesIO(res.body))["TOTAL"]["C3"].value == TODAY.isoformat()

"""IBM/레노버 x86 문서 판별이 화면 없이도 코어까지 닿는지.

화면에는 더 이상 고르는 토글이 없다. 업로드된 XML 내용으로 알아낸
읽기 방식(`quotation.core.modes.detect`)은 견적 내용이 아니라 진단
정보이므로 구조화 로그에 남긴다 (계획서 §13).
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import load_workbook

import api

TODAY = dt.date(2026, 8, 19)


def _convert(fixtures, template_bytes, name="integrated_quote.xml"):
    upload = api.Upload(filename=name, content=(fixtures / name).read_bytes(),
                        content_type="text/xml")
    return api.convert_response(
        [upload], template_bytes=template_bytes, template_version="v-test",
        deployment_version="d-test", request_id="req-1", today=TODAY)


def _sheets(response) -> list[str]:
    return load_workbook(BytesIO(response.body)).sheetnames


# --- 문서 내용으로 갈리는지 ------------------------------------------------------

def test_ibm_document_takes_the_unix_path(fixtures, template_bytes):
    """ProductName 이 없는 문서(IBM)는 지금까지와 같이 돈다."""
    res = _convert(fixtures, template_bytes, name="new_quote.xml")
    assert res.status == 200
    assert res.log["mode"] == "unix"


def test_lenovo_document_takes_the_integrated_path(fixtures, template_bytes):
    """ProductName 이 있는 문서(레노버 x86)는 자동으로 통합 모드로 읽힌다."""
    res = _convert(fixtures, template_bytes, name="integrated_quote.xml")
    assert res.status == 200
    assert res.log["mode"] == "integrated"
    assert _sheets(res) == ["TOTAL", "백업서버_1식", "메일-스펨_2식",
                            "SAMPLE 42U DEEP STATIC RACK", "template"]


def test_integrated_document_does_not_double_count(fixtures, template_bytes):
    total = load_workbook(BytesIO(
        _convert(fixtures, template_bytes, name="integrated_quote.xml").body))["TOTAL"]
    assert total["G8"].value == 40172000   # 본체 LP 하나가 그 서버의 전체 금액
    assert total["G9"].value is None       # S/W·서비스는 금액을 더하지 않는다


# --- 배선 (화면에는 더 이상 이 개념이 없다) --------------------------------------

def test_the_screen_offers_no_mode_toggle():
    """토글을 없앴다 — 화면 어디에도 `name="mode"` 가 없어야 한다."""
    from pathlib import Path

    html = (Path(__file__).resolve().parents[1] / "frontend" / "index.html"
            ).read_text(encoding="utf-8")
    assert 'name="mode"' not in html

"""골든의 테두리/채우기/정렬 패턴을 압축 표로 출력. 장식 규칙 도출용.

테두리 표기: 왼오위아래 순서, m=medium t=thin d=double h=hair .=없음
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ABBR = {None: ".", "medium": "m", "thin": "t", "double": "d", "hair": "h",
        "dotted": "o", "dashed": "s", "thick": "k"}
COLS = "BCDEFGHI"


def border(c):
    b = c.border
    return "".join(ABBR.get(getattr(getattr(b, s), "style", None), "?")
                   for s in ("left", "right", "top", "bottom"))


def fill(c):
    f = c.fill
    if f is None or f.patternType is None:
        return "-"
    try:
        rgb = f.fgColor.rgb
        return rgb if isinstance(rgb, str) else f"idx{f.fgColor.indexed}"
    except (TypeError, ValueError):
        return "?"


def valign(c):
    a = c.alignment
    v = (a.vertical or "")[:3]
    return v + ("W" if a.wrap_text else "")


def fontcolor(c):
    col = c.font.color
    if col is None:
        return "-"
    try:
        return col.rgb if isinstance(col.rgb, str) else f"t{col.theme}"
    except (TypeError, ValueError):
        return "?"


def dump(path, sheet, rows):
    ws = load_workbook(path)[sheet]
    print(f"=== [{sheet}]  {Path(path).name}")
    for label, fn in (("테두리", border), ("채우기", fill),
                      ("세로/줄바꿈", valign), ("글꼴색", fontcolor)):
        print(f"--- {label}")
        print("행  높이 | " + " | ".join(f"{c:^8}" for c in COLS))
        for r in rows:
            h = ws.row_dimensions[r].height
            hs = f"{h:>4}" if h else "   -"
            cells = " | ".join(f"{fn(ws[f'{c}{r}']):^8}" for c in COLS)
            val = str(ws[f"B{r}"].value or ws[f"C{r}"].value or "")[:18]
            print(f"{r:<3} {hs} | {cells}  {val}")
        print()


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    gold = root / ".cache" / "FS5045_260722.xlsx"
    which = sys.argv[1] if len(sys.argv) > 1 else "detail"
    if which == "total":
        dump(gold, "TOTAL", range(5, 24))
    else:
        dump(gold, "4680-3P4 #1", range(5, 32))

"""템플릿 두 판본의 셀 차이 비교 (글꼴·서식·값)."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _color(c):
    if c is None:
        return None
    try:
        return c.rgb if isinstance(c.rgb, str) else f"idx{c.indexed}"
    except (TypeError, ValueError):
        return "?"


def facts(c):
    """셀의 겉모습 전부. 글꼴만 보면 채우기·테두리 변경을 놓친다."""
    f, fill, b = c.font, c.fill, c.border
    return {
        "값": c.value,
        "글꼴": f.name,
        "크기": f.size,
        "볼드": bool(f.b),
        "글꼴색": _color(f.color),
        "서식": c.number_format,
        "가로": c.alignment.horizontal,
        "세로": c.alignment.vertical,
        "줄바꿈": bool(c.alignment.wrap_text),
        "채우기": (fill.patternType, _color(fill.fgColor)) if fill else None,
        "테두리": tuple(getattr(getattr(b, s), "style", None)
                     for s in ("left", "right", "top", "bottom")),
    }


def main(old_path, new_path, max_row=40, max_col=10):
    old = load_workbook(old_path)
    new = load_workbook(new_path)
    print(f"시트: 이전={old.sheetnames} 이후={new.sheetnames}")
    for name in new.sheetnames:
        if name not in old.sheetnames:
            print(f"[{name}] 새 시트")
            continue
        o, n = old[name], new[name]
        print(f"\n=== [{name}]")
        found = 0
        for r in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                addr = f"{get_column_letter(col)}{r}"
                fo, fn = facts(o[addr]), facts(n[addr])
                if fo == fn:
                    continue
                changed = {k: (fo[k], fn[k]) for k in fo if fo[k] != fn[k]}
                print(f"  {addr}: " + ", ".join(
                    f"{k} {a!r} -> {b!r}" for k, (a, b) in changed.items()))
                found += 1
        if not found:
            print("  (차이 없음)")


if __name__ == "__main__":
    root = Path(__file__).resolve().parents[1]
    main(sys.argv[1] if len(sys.argv) > 1 else root / ".cache" / "견적서_template.OLD.xlsx",
         sys.argv[2] if len(sys.argv) > 2 else root / ".cache" / "견적서_template.xlsx")

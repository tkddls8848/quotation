"""회귀 비교기 — 생성물이 골든 견적서와 동일한지 셀 단위로 검증.

두 .xlsx 를 비교한다. 골든이 .xls 라면 tools/xls2xlsx.ps1 로 먼저 변환할 것.

비교 항목 (SPEC_CELLMAP.md §6):
  1. 시트 목록/순서/숨김 상태
  2. 셀 수식 문자열
  3. 셀 값
  4. 숫자 서식 / 가로 정렬 / Bold
  5. 병합 영역
  6. 열 너비

사용법:
    python tools/compare.py <golden.xlsx> <actual.xlsx> [--ignore-file ignore.txt]
종료 코드: 0 = 일치, 1 = 차이 있음
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 비교할 열 범위 (템플릿 사용 범위는 B:I, 여유분 포함)
MAX_COL = 12
MAX_ROW = 200


@dataclass
class Diff:
    sheet: str
    cell: str
    attr: str
    expected: object
    actual: object

    def __str__(self) -> str:
        return (f"[{self.sheet}] {self.cell} {self.attr}: "
                f"골든={self.expected!r} / 생성={self.actual!r}")


@dataclass
class Report:
    diffs: list[Diff] = field(default_factory=list)
    ignored: int = 0

    def add(self, *a):
        self.diffs.append(Diff(*a))

    @property
    def ok(self) -> bool:
        return not self.diffs


def _norm_num(v):
    """숫자 비교 정규화. 1 과 1.0, Decimal 과 float 을 동일 취급."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        f = float(v)
        return int(f) if f.is_integer() else round(f, 9)
    return v


def _norm(v):
    if v is None:
        return None
    v = _norm_num(v)
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def _norm_fmt(f):
    """숫자 서식 정규화. 'G/표준' 과 'General' 은 같은 것."""
    if f in (None, "", "General", "G/표준"):
        return "General"
    return f


def _color(c):
    """Color -> 비교 가능한 튜플. rgb/theme/indexed 를 구분해 읽는다."""
    if c is None:
        return None
    kind = getattr(c, "type", None)
    try:
        if kind == "rgb":
            return ("rgb", c.rgb)
        if kind == "theme":
            return ("theme", c.theme, c.tint)
        if kind == "indexed":
            return ("indexed", c.indexed)
    except (TypeError, ValueError):
        return (kind, "?")
    return (kind,)


def _fill(cell):
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    return (f.patternType, _color(f.fgColor), _color(f.bgColor))


def _border(cell):
    b = cell.border
    out = []
    for side in ("left", "right", "top", "bottom"):
        s = getattr(b, side, None)
        out.append((getattr(s, "style", None), _color(getattr(s, "color", None))))
    return tuple(out)


def compare_sheet(name, gs, as_, rep: Report, ignore: set[str]):
    # 병합 영역
    g_merges = {str(r) for r in gs.merged_cells.ranges}
    a_merges = {str(r) for r in as_.merged_cells.ranges}
    for m in sorted(g_merges - a_merges):
        rep.add(name, m, "병합 누락", m, None)
    for m in sorted(a_merges - g_merges):
        rep.add(name, m, "병합 초과", None, m)

    # 열 너비
    for c in range(1, MAX_COL + 1):
        letter = get_column_letter(c)
        gw = gs.column_dimensions[letter].width if letter in gs.column_dimensions else None
        aw = as_.column_dimensions[letter].width if letter in as_.column_dimensions else None
        if gw is not None and aw is not None:
            if abs(gw - aw) > 0.05:
                rep.add(name, f"{letter}:{letter}", "열너비", round(gw, 2), round(aw, 2))
        elif (gw is None) != (aw is None):
            rep.add(name, f"{letter}:{letter}", "열너비", gw, aw)

    # 행 높이
    for row in range(1, MAX_ROW + 1):
        gh = gs.row_dimensions[row].height if row in gs.row_dimensions else None
        ah = as_.row_dimensions[row].height if row in as_.row_dimensions else None
        if gh != ah:
            rep.add(name, f"{row}행", "행높이", gh, ah)

    # 셀
    for row in range(1, MAX_ROW + 1):
        for col in range(1, MAX_COL + 1):
            addr = f"{get_column_letter(col)}{row}"
            if f"{name}!{addr}" in ignore:
                rep.ignored += 1
                continue
            g = gs.cell(row=row, column=col)
            a = as_.cell(row=row, column=col)

            gv, av = _norm(g.value), _norm(a.value)
            if gv != av:
                rep.add(name, addr, "값/수식", gv, av)
                continue  # 값이 다르면 서식 비교는 노이즈

            # 테두리·채우기는 값이 없는 셀에도 그려진다 (표 괘선).
            if _border(g) != _border(a):
                rep.add(name, addr, "테두리", _border(g), _border(a))
            if _fill(g) != _fill(a):
                rep.add(name, addr, "채우기", _fill(g), _fill(a))

            if gv is None:
                continue

            gf, af = _norm_fmt(g.number_format), _norm_fmt(a.number_format)
            if gf != af:
                rep.add(name, addr, "숫자서식", gf, af)

            ga, aa = g.alignment, a.alignment
            if (ga.horizontal or "general") != (aa.horizontal or "general"):
                rep.add(name, addr, "가로정렬", ga.horizontal, aa.horizontal)
            if (ga.vertical or "bottom") != (aa.vertical or "bottom"):
                rep.add(name, addr, "세로정렬", ga.vertical, aa.vertical)
            if bool(ga.wrap_text) != bool(aa.wrap_text):
                rep.add(name, addr, "줄바꿈", bool(ga.wrap_text),
                        bool(aa.wrap_text))

            if bool(g.font.bold) != bool(a.font.bold):
                rep.add(name, addr, "Bold", bool(g.font.bold), bool(a.font.bold))

            if g.font.name != a.font.name:
                rep.add(name, addr, "글꼴", g.font.name, a.font.name)

            if float(g.font.size or 0) != float(a.font.size or 0):
                rep.add(name, addr, "글꼴크기", g.font.size, a.font.size)

            if _color(g.font.color) != _color(a.font.color):
                rep.add(name, addr, "글꼴색", _color(g.font.color),
                        _color(a.font.color))


def compare(golden_path, actual_path, ignore: set[str]) -> Report:
    rep = Report()
    gw = load_workbook(golden_path, data_only=False)
    aw = load_workbook(actual_path, data_only=False)

    if gw.sheetnames != aw.sheetnames:
        rep.add("<workbook>", "-", "시트 목록/순서", gw.sheetnames, aw.sheetnames)
        common = [s for s in gw.sheetnames if s in aw.sheetnames]
    else:
        common = gw.sheetnames

    for name in common:
        gs, as_ = gw[name], aw[name]
        if gs.sheet_state != as_.sheet_state:
            rep.add(name, "-", "시트 표시상태", gs.sheet_state, as_.sheet_state)
        if (gs.print_area or None) != (as_.print_area or None):
            rep.add(name, "-", "인쇄영역", gs.print_area, as_.print_area)
        compare_sheet(name, gs, as_, rep, ignore)

    return rep


def load_ignore(path) -> set[str]:
    """무시 목록. 한 줄에 'SHEET!ADDR' 형식. '#' 주석 허용."""
    if not path:
        return set()
    out = set()
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.split("#", 1)[0].strip()
            if line:
                out.add(line)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("golden")
    ap.add_argument("actual")
    ap.add_argument("--ignore-file")
    ap.add_argument("--max-print", type=int, default=60)
    args = ap.parse_args()

    rep = compare(args.golden, args.actual, load_ignore(args.ignore_file))

    if rep.ok:
        print(f"OK  차이 없음  (무시 {rep.ignored}건)")
        return 0

    print(f"FAIL  차이 {len(rep.diffs)}건  (무시 {rep.ignored}건)\n")
    for d in rep.diffs[:args.max_print]:
        print("  " + str(d))
    if len(rep.diffs) > args.max_print:
        print(f"  … 외 {len(rep.diffs) - args.max_print}건")
    return 1


if __name__ == "__main__":
    sys.exit(main())

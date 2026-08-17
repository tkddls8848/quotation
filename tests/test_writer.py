"""Phase 3 회귀 — 생성물이 골든 견적서와 셀 단위로 같은지 검증."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]

from quotation.core import xml_reader  # noqa: E402
from quotation.core.writer import ibm_writer  # noqa: E402

import compare as cmp_mod  # noqa: E402

SAMPLES = ROOT / "samples"
CACHE = ROOT / "cache" if (ROOT / "cache").exists() else ROOT / ".cache"
TEMPLATE = CACHE / "견적서_template.xlsx"

#: (골든 이름, 그 골든을 만든 날짜). C3 에 그날 날짜가 들어가므로 맞춰야 한다.
CASES = [
    ("FS5045_260722", dt.date(2026, 7, 23)),
    ("X-ROIS 통합서버#2", dt.date(2026, 7, 23)),
    ("1080MES", dt.date(2026, 7, 23)),  # 증설 견적 (BASE/PROPOSED/UPGRADE/REMOVE)
    ("FS9500_2TB_19.2TB 12ea_U150TB_16port (수익영역 스토리지 Active)_251223",
     dt.date(2025, 12, 23)),  # 장비군 1개, 이름·괄호·공백이 섞인 화일명
    ("TS4300_FC LTO9 6ea_LTO9 40ea_cleaning 5ea_구성파일_260706",
     dt.date(2026, 7, 23)),  # 'No CPUSIU' 그룹, H/W 블록 1개 구간
]

pytestmark = pytest.mark.skipif(
    not TEMPLATE.exists(),
    reason="tools/xls2xlsx.ps1 로 .cache 를 먼저 생성하십시오",
)


def _build(name: str, today: dt.date, tmp_path: Path) -> Path:
    source = SAMPLES / f"{name}.xml"
    if not source.exists():
        pytest.skip(f"샘플 없음: {source.name}")
    quote = xml_reader.parse(source)
    return ibm_writer.write(quote, TEMPLATE, tmp_path / f"{name}.xlsx", today=today)


#: writer 가 직접 손대는 머리말 셀. 나머지 1~7행은 템플릿에서 그대로 와야 한다.
#: 상세 시트의 H7 은 값을 지우고 H6:H7 로 병합하는 자리다 (SPEC_CELLMAP.md §4.1).
WRITTEN_HEADER_CELLS = {"TOTAL": {"B2", "C3"}, "detail": {"C1", "C3", "H7"}}


#: 대조할 골든이 없는 샘플. 최소한 오류 없이 변환되는지는 지킨다.
SMOKE_ONLY = ["운영 DB서버 #1_251223"]


@pytest.mark.parametrize("name", SMOKE_ONLY)
def test_converts_without_error(name, tmp_path):
    """골든이 없어 값 대조는 못 하지만 변환 자체는 깨지면 안 된다."""
    from quotation.core import convert

    source = SAMPLES / f"{name}.xml"
    if not source.exists():
        pytest.skip(f"샘플 없음: {source.name}")
    xml = tmp_path / source.name
    xml.write_bytes(source.read_bytes())

    result = convert.convert(xml, today=dt.date(2025, 12, 23))
    assert result.output.is_file()
    assert result.group_count > 0
    # TOTAL + 상세 + 숨김 template
    assert len(load_workbook(result.output).sheetnames) == result.group_count + 2


def test_template_header_passes_through(tmp_path):
    """템플릿 1~7행의 글꼴·서식이 산출물에 그대로 전달되는지.

    템플릿 글꼴을 고치면 결과에 바로 반영되어야 한다. writer 가 값을 쓰는
    몇 셀만 예외다. 이 테스트가 있어야 템플릿 수정이 조용히 무시되지 않는다.
    """
    from openpyxl.utils import get_column_letter

    actual = _build("FS5045_260722", dt.date(2026, 7, 23), tmp_path)
    tw = load_workbook(TEMPLATE)
    aw = load_workbook(actual)

    for sheet, source in (("TOTAL", "TOTAL"), ("4680-3P4 #1", "template")):
        written = WRITTEN_HEADER_CELLS["TOTAL" if sheet == "TOTAL" else "detail"]
        t, a = tw[source], aw[sheet]
        for row in range(1, 8):
            for col in range(2, 9):
                addr = f"{get_column_letter(col)}{row}"
                if addr in written:
                    continue
                tc, ac = t[addr], a[addr]
                if tc.value is None:
                    continue
                assert (tc.font.name, tc.font.size) == (ac.font.name, ac.font.size), (
                    f"[{sheet}] {addr} 글꼴이 템플릿과 다르다"
                )
                assert tc.value == ac.value, f"[{sheet}] {addr} 값이 템플릿과 다르다"


def test_total_sheet_keeps_template_drawings(tmp_path):
    """TOTAL 시트 상단의 로고·머리글 도형이 남아 있어야 한다.

    openpyxl 은 저장 시 도형을 버린다. 이것이 빠지면 첫 장 상단이 비어 보인다.
    상세 시트에는 골든과 마찬가지로 그림이 없어야 한다.

    숨김 template 시트까지 비었는지는 Pillow 유무에 따라 갈리므로
    `tests/test_bytes_api.py` 가 따로 본다.
    """
    import re
    import zipfile

    actual = _build("FS5045_260722", dt.date(2026, 7, 23), tmp_path)
    with zipfile.ZipFile(actual) as z:
        names = z.namelist()
        media = [n for n in names if n.startswith("xl/media/")]
        drawn = sorted(n for n in names
                       if n.startswith("xl/worksheets/sheet")
                       and re.search(rb"<drawing\s", z.read(n)))
        shapes = sum(z.read(n).count(b"<xdr:sp ") for n in names
                     if n.startswith("xl/drawings/drawing"))

    assert media, "템플릿의 로고 이미지가 옮겨지지 않았다"
    assert "xl/worksheets/sheet1.xml" in drawn, f"TOTAL(sheet1)에 그림이 없다: {drawn}"
    assert shapes, "머리글 도형이 옮겨지지 않았다 (openpyxl 은 도형을 만들지 못한다)"
    assert "xl/worksheets/sheet2.xml" not in drawn, f"상세 시트에 그림이 남았다: {drawn}"


@pytest.mark.parametrize("name,today", CASES)
def test_matches_golden(name, today, tmp_path):
    golden = CACHE / f"{name}.xlsx"
    if not golden.exists():
        pytest.skip(f"골든 없음: {golden}")

    actual = _build(name, today, tmp_path)
    ignore = cmp_mod.load_ignore(ROOT / "tests" / "golden_ignore.txt")
    rep = cmp_mod.compare(golden, actual, ignore)

    if not rep.ok:
        lines = "\n".join("  " + str(d) for d in rep.diffs[:40])
        extra = "" if len(rep.diffs) <= 40 else f"\n  … 외 {len(rep.diffs) - 40}건"
        pytest.fail(f"{name}: 골든과 차이 {len(rep.diffs)}건\n{lines}{extra}")

"""변환 오케스트레이션, 설정, GUI 인수 동작."""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from quotation import config as config_mod, paths  # noqa: E402
from quotation.core import convert, xml_reader  # noqa: E402
from quotation.core.xml_reader import QuotationXmlError  # noqa: E402

SAMPLES = ROOT / "samples"
FS5045 = SAMPLES / "FS5045_260722.xml"
XROIS = SAMPLES / "X-ROIS 통합서버#2.xml"
TODAY = dt.date(2026, 7, 23)


def _require_sample(source: Path) -> None:
    if not source.exists():
        pytest.skip(f"샘플 없음: {source.name}")


# --- 리소스 ------------------------------------------------------------------

def test_template_ships_with_package():
    assert paths.template_path().is_file(), "템플릿이 패키지에 포함되어야 한다"


# --- 변환 --------------------------------------------------------------------

def _convert(tmp_path, source: Path = FS5045, **kwargs):
    """샘플을 tmp_path 로 복사해 변환한다. 산출물은 그 옆에 생긴다."""
    _require_sample(source)
    xml = tmp_path / source.name
    xml.write_bytes(source.read_bytes())
    return convert.convert(xml, today=TODAY, **kwargs)


def test_convert_writes_next_to_xml(tmp_path):
    _require_sample(FS5045)
    xml = tmp_path / "sample.xml"
    xml.write_bytes(FS5045.read_bytes())
    result = convert.convert(xml, today=TODAY)
    assert result.output == tmp_path / "sample.xlsx"
    assert result.output.is_file()
    assert result.group_count == 3


def test_output_location_is_fixed(tmp_path):
    """저장 위치는 언제나 XML 과 같은 폴더다. 바꿀 수단이 없다."""
    nested = tmp_path / "깊은" / "폴더"
    nested.mkdir(parents=True)
    result = _convert(nested)
    assert result.output.parent == nested
    assert convert.output_path_for(nested / "a.xml") == nested / "a.xlsx"


def test_progress_reaches_100(tmp_path):
    seen: list[tuple[int, str]] = []
    _convert(tmp_path, progress=lambda p, m: seen.append((p, m)))
    assert seen[0][0] < seen[-1][0]
    assert seen[-1][0] == 100
    # 상태 메시지는 원본 프로그램의 문구를 유지한다
    assert seen[-1][1] == "견적서작성을 완료하였습니다."


def test_bad_xml_reports_original_message(tmp_path):
    bad = tmp_path / "bad.xml"
    bad.write_text("<CFXML><CFData/></CFXML>", encoding="utf-8")
    with pytest.raises(QuotationXmlError, match="Item을 찾을 수 없습니다"):
        convert.convert(bad, today=TODAY)


def test_conversion_preserves_component_quantity_and_lp(tmp_path):
    xml = tmp_path / "sample.xml"
    xml.write_text("""<?xml version="1.0" encoding="UTF-8"?>
<CFXML><CFData><ProductLineItem>
  <ProductLineNumber>1000</ProductLineNumber>
  <CPUSIUvalue>1</CPUSIUvalue>
  <TransactionType>NEW</TransactionType>
  <ProprietaryGroupIdentifier>1000</ProprietaryGroupIdentifier>
  <Quantity>2</Quantity>
  <ProductIdentification><PartnerProductIdentification>
    <ProductDescription>IBM Test System</ProductDescription>
    <ProductTypeCode>Hardware</ProductTypeCode>
    <ProprietaryProductIdentifier>TEST-001</ProprietaryProductIdentifier>
  </PartnerProductIdentification></ProductIdentification>
  <UnitListPrice><FinancialAmount><MonetaryAmount>1,234.50</MonetaryAmount></FinancialAmount></UnitListPrice>
  <ProductSubLineItem>
    <TransactionType>ADD</TransactionType><Quantity>3</Quantity>
    <ProductIdentification><PartnerProductIdentification>
      <ProductDescription>Test Component</ProductDescription>
      <ProductTypeCode>Hardware</ProductTypeCode>
      <ProprietaryProductIdentifier>PART-001</ProprietaryProductIdentifier>
    </PartnerProductIdentification></ProductIdentification>
    <UnitListPrice><FinancialAmount><MonetaryAmount>100</MonetaryAmount></FinancialAmount></UnitListPrice>
  </ProductSubLineItem>
</ProductLineItem></CFData></CFXML>""", encoding="utf-8")

    result = convert.convert(xml, today=TODAY,
                             template=paths.resource_dir() / paths.TEMPLATE_NAME)
    ws = load_workbook(result.output, data_only=False)["TEST SYSTEM"]

    assert [ws[f"{col}8"].value for col in "CDEFG"] == [
        "TEST-001", "IBM Test System", 2, 1234.5, "=E8*F8",
    ]
    assert [ws[f"{col}9"].value for col in "CDEFG"] == [
        "PART-001", "Test Component", "=3*E8", 100, "=E9*F9",
    ]


# --- 옵션 --------------------------------------------------------------------

def test_maintenance_is_never_written(tmp_path):
    """유지정비료를 견적서 H·I열에 쓰지 않는다."""
    out = _convert(tmp_path, XROIS).output
    wb = load_workbook(out)

    server = wb["SERVER 1"]
    assert server["H51"].value is None
    assert server["I51"].value is None
    assert wb["TOTAL"]["H10"].value is None

    # H열 어디에도 값이나 합계 수식이 없어야 한다
    for sheet in wb.worksheets:
        for row in range(1, 120):
            for col in ("H", "I"):
                value = sheet[f"{col}{row}"].value
                assert value is None or row <= 7, (
                    f"[{sheet.title}] {col}{row} 에 유지정비료가 남아 있다: {value!r}")


def _quote_number(tmp_path, template_b2: str) -> str:
    from openpyxl import load_workbook as lw

    from quotation.core.writer import ibm_writer

    _require_sample(FS5045)
    template = tmp_path / "t.xlsx"
    wb = lw(paths.template_path())
    wb["TOTAL"]["B2"] = template_b2
    wb.save(template)

    quote = xml_reader.parse(FS5045)
    out = ibm_writer.write(quote, template, tmp_path / "o.xlsx", today=TODAY)
    return lw(out)["TOTAL"]["B2"].value


def test_quote_number_updates_year_only(tmp_path):
    """결과는 언제나 'NO : Trialinfo-{YY}-' 다. 연도 뒤에 아무것도 붙지 않는다."""
    assert _quote_number(tmp_path, "NO : Trialinfo-24-") == "NO : Trialinfo-26-"


def test_quote_number_keeps_custom_prefix(tmp_path):
    """앞부분 문구는 템플릿에 적힌 그대로 둔다."""
    assert _quote_number(tmp_path, "견적 NO : Trialinfo-20-") == "견적 NO : Trialinfo-26-"


def test_template_lives_outside_the_exe(tmp_path, monkeypatch):
    """템플릿은 EXE 옆의 고칠 수 있는 파일이어야 한다.

    없으면 번들 사본으로 한 번 만들어 준다. 안에 묻어 두면 견적서 번호와
    담당자 이름을 고칠 수 없다.
    """
    monkeypatch.setattr(paths, "app_dir", lambda: tmp_path)

    target = tmp_path / paths.TEMPLATE_NAME
    assert not target.exists()
    assert paths.template_path() == target
    assert target.exists(), "번들 사본으로 자동 생성되어야 한다"

    # 이미 있으면 덮어쓰지 않는다 (사용자 편집 보존)
    target.write_bytes(b"edited")
    assert paths.template_path().read_bytes() == b"edited"

def test_main_opens_gui_with_first_file_argument(monkeypatch):
    from quotation import __main__
    from quotation.ui import main_window

    seen = []
    monkeypatch.setattr(main_window, "run", lambda prefill: seen.append(prefill) or 0)

    assert __main__.main(["sample.xml", "ignored.xml"]) == 0
    assert seen == ["sample.xml"]


def test_supply_row_is_always_blank(tmp_path):
    """공급가 행은 언제나 공란이다. 골든 전부 그러하고 할인 기능은 없앴다."""
    wb = load_workbook(_convert(tmp_path).output)
    assert wb["TOTAL"]["G21"].value is None
    assert wb["4680-3P4 #1"]["G29"].value is None


# --- 설정 --------------------------------------------------------------------

def test_config_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setattr(paths, "app_data_dir", lambda: tmp_path)
    monkeypatch.setattr(paths, "config_path", lambda: tmp_path / "config.json")

    cfg = config_mod.load()
    cfg.open_result_when_done = False
    cfg.remember(Path(r"C:\in\a.xml"))
    config_mod.save(cfg)

    again = config_mod.load()
    assert again.open_result_when_done is False
    assert again.last_input_dir == r"C:\in"

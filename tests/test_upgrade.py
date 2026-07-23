"""증설 견적 — 참조 구성 제외와 제거 부품 표시.

골든: samples/1080MES.{xml,xls}
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from quotation.core import convert, xml_reader  # noqa: E402
from quotation.core.writer.decorate import RED  # noqa: E402

SAMPLE = ROOT / "samples" / "1080MES.xml"
TODAY = dt.date(2026, 7, 23)


@pytest.fixture(scope="module")
def quote():
    return xml_reader.parse(SAMPLE)


@pytest.fixture(scope="module")
def book(tmp_path_factory):
    # 산출물은 XML 옆에 생기므로 샘플을 임시 폴더로 복사해 변환한다
    work = tmp_path_factory.mktemp("up")
    xml = work / SAMPLE.name
    xml.write_bytes(SAMPLE.read_bytes())
    return load_workbook(convert.convert(xml, today=TODAY).output)


# --- 참조 구성 제외 -------------------------------------------------------------

def test_base_and_proposed_are_excluded(quote):
    """기존(BASE)·증설후(PROPOSED) 구성은 견적서에 넣지 않는다.

    XML 29줄 중 BASE 11 + PROPOSED 13 을 빼면 증설분 5줄만 남는다.
    """
    kinds = {i.txn_type for g in quote.groups for i in g.items}
    assert kinds == {"UPGRADE", "DISCO", "NEW"}
    assert sum(len(g.items) for g in quote.groups) == 5


def test_upgrade_becomes_one_equipment_group(quote):
    """본체 라인이 없는 그룹(DISCO/NEW)은 앞 그룹에 붙어 한 장이 된다."""
    assert len(quote.groups) == 1
    assert quote.groups[0].sheet_name == "SERVER 1"


def test_equipment_name_comes_from_base_config(quote):
    """증설 라인의 설명에는 장비 이름이 없다. 이름은 BASE 구성에서 딴다.

    UPGRADE 라인 설명은 '9080 Model HEU' 지만 장비 이름은 'Server 1' 이다.
    """
    assert quote.groups[0].item_key == "Server 1"
    assert quote.groups[0].items[0].description == "9080 Model HEU"


def test_only_one_detail_sheet(book):
    assert book.sheetnames == ["TOTAL", "SERVER 1", "template"]


def test_base_items_absent_from_sheet(book):
    """기존 구성에만 있던 품목이 견적서에 나오면 안 된다."""
    ws = book["SERVER 1"]
    models = {ws[f"C{r}"].value for r in range(8, 60)}
    assert "9080-HEX" not in models, "기존 장비(BASE)가 견적서에 남아 있다"
    assert "9080-HEX_9080-HEU" in models, "증설(UPGRADE) 라인은 있어야 한다"


# --- 제거 부품 ----------------------------------------------------------------

REMOVE_ROWS_UPGRADE = (19, 20, 21)   # EBAB, EDFP, ESWK
REMOVE_ROWS_DISCO = (34, 35, 36, 37, 38, 39)


@pytest.mark.parametrize("row,expected", [
    (19, "=-1*E8"), (20, "=-2*E8"), (21, "=-39*E8"),
])
def test_removed_parts_have_negative_quantity(book, row, expected):
    """제거 부품은 수량을 음수로 적는다. XML 의 수량은 양수로 들어온다."""
    assert book["SERVER 1"][f"E{row}"].value == expected


def test_removed_parts_without_priced_parent(book):
    """기준행에 가격이 없으면 부모 참조 없이 '=-1' 로 적는다 (DISCO 블록)."""
    ws = book["SERVER 1"]
    for row in REMOVE_ROWS_DISCO:
        assert ws[f"E{row}"].value == "=-1"


@pytest.mark.parametrize("row", REMOVE_ROWS_UPGRADE + REMOVE_ROWS_DISCO)
def test_removed_parts_are_red(book, row):
    ws = book["SERVER 1"]
    for col in "CDEFG":
        color = ws[f"{col}{row}"].font.color
        assert color is not None and color.rgb == RED, f"{col}{row} 이 붉지 않다"


@pytest.mark.parametrize("row", REMOVE_ROWS_UPGRADE + REMOVE_ROWS_DISCO)
def test_removed_parts_have_no_amount(book, row):
    """제거 부품은 단가·금액을 비운다."""
    ws = book["SERVER 1"]
    assert ws[f"F{row}"].value is None
    assert ws[f"G{row}"].value is None


def test_normal_rows_stay_uncoloured(book):
    ws = book["SERVER 1"]
    for row in (8, 10, 22, 25):
        assert ws[f"C{row}"].font.color is None, f"{row}행이 물들었다"


# --- 합계 ---------------------------------------------------------------------

def test_total_sheet_uses_sum_even_for_one_group(book):
    """TOTAL 은 그룹이 하나여도 =SUM(...) 을 쓴다. 상세 시트와 다르다."""
    assert book["TOTAL"]["G14"].value == "=SUM(G13)"


def test_detail_grand_total_enumerates_sections(book):
    assert book["SERVER 1"]["G52"].value == "=SUM(G41,G51)"


def test_amounts_match_golden(book):
    """골든 TOTAL: H/W 소계 2,621,454.9 / S/W 소계 271,750.8."""
    ws = book["TOTAL"]
    assert ws["G8"].value == 2621454.9
    assert ws["G10"].value == 271750.8
    assert ws["G13"].value == "=SUM(G8:G12)"

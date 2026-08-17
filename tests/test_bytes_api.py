"""바이트 I/O 코어 — 웹 Worker 가 쓰는 순수 경로.

계약은 하나다. **같은 XML 이면 경로 방식과 바이트 방식의 결과가 의미상 같다.**
데스크톱과 웹이 한동안 함께 운영되므로(계획서 §11 Phase 6) 이 동등성이 깨지면
두 경로의 견적서가 달라진다.
"""
from __future__ import annotations

import datetime as dt
import importlib.util
import re
import zipfile
from io import BytesIO

import pytest

import compare as cmp_mod
from quotation.core import convert, xml_reader
from quotation.core.writer import drawings, ibm_writer

TODAY = dt.date(2026, 7, 23)

FIXTURE_NAMES = ["new_quote.xml", "upgrade_quote.xml", "no_charge.xml",
                 "euckr_quote.xml"]

#: openpyxl 은 Pillow 가 있을 때만 워크북의 **이미지**를 읽어 다시 써 준다.
#: 배포되는 세 갈래 어디에도 Pillow 는 없다 — 데스크톱 EXE(requirements.txt),
#: Worker(web/pyproject.toml), 브라우저 엔진(build_browser_engine.py 의 wheel
#: 목록) 모두 lxml·openpyxl·et_xmlfile 뿐이다. 그래서 운영에서는 그림이 전부
#: 버려지고 `drawings.carry_over_bytes` 가 TOTAL 것만 되살린다.
#:
#: 개발 기계에 Pillow 가 딸려 들어와 있으면 openpyxl 이 숨김 template 시트의
#: 로고까지 다시 써서 산출물이 한 장 더 그려진다. 도형(텍스트박스·사각형)은
#: 어느 쪽이든 openpyxl 이 살리지 못하므로, 아래 검사는 도형을 기준으로 삼고
#: Pillow 에 따라 갈리는 부분만 따로 떼어 둔다.
HAS_PILLOW = importlib.util.find_spec("PIL") is not None


def _without_drawings(xlsx: bytes) -> bytes:
    """그림·이미지를 걷어낸 사본. '그림 없는 양식' 을 만들어 준다.

    openpyxl 이 그림을 버린다는 성질에 기대면 Pillow 유무에 따라 결과가 갈린다.
    검사에 쓸 입력은 여기서 직접 만든다.
    """
    out = BytesIO()
    with zipfile.ZipFile(BytesIO(xlsx)) as src, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
        for name in src.namelist():
            if name.startswith(("xl/drawings/", "xl/media/")):
                continue
            body = src.read(name)
            if name.endswith(".rels"):
                body = re.sub(
                    rb"<Relationship\b[^>]*?(?:drawings|media)[^>]*?/>", b"", body)
            elif name == "[Content_Types].xml":
                body = re.sub(rb"<Override\b[^>]*?/xl/drawings/[^>]*?/>", b"", body)
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                body = re.sub(rb"<drawing\b[^>]*?/>", b"", body)
            dst.writestr(name, body)
    return out.getvalue()


def _shape_count(xlsx: bytes) -> int:
    """그림 부품에 담긴 도형(<xdr:sp>) 개수. openpyxl 은 이것을 만들지 못한다."""
    with zipfile.ZipFile(BytesIO(xlsx)) as z:
        return sum(z.read(n).count(b"<xdr:sp ") for n in z.namelist()
                   if n.startswith("xl/drawings/drawing"))


def _sheets_with_drawings(xlsx: bytes) -> list[str]:
    with zipfile.ZipFile(BytesIO(xlsx)) as z:
        return sorted(n for n in z.namelist()
                      if n.startswith("xl/worksheets/sheet")
                      and re.search(rb"<drawing\s", z.read(n)))


# --- 파싱 --------------------------------------------------------------------

@pytest.mark.parametrize("name", FIXTURE_NAMES)
def test_parse_bytes_matches_parse_path(name, fixtures):
    source = fixtures / name
    assert xml_reader.parse_bytes(source.read_bytes()) == xml_reader.parse(source)


def test_parse_bytes_rejects_empty_input():
    with pytest.raises(xml_reader.QuotationXmlError, match="빈 화일"):
        xml_reader.parse_bytes(b"   \n")


def test_parse_bytes_reports_malformed_xml():
    with pytest.raises(xml_reader.QuotationXmlError, match="장애 발생"):
        xml_reader.parse_bytes(b"<CFXML><CFData>")


def test_parse_bytes_blocks_external_entities(tmp_path):
    """외부 엔티티가 확장되면 안 된다 (XXE)."""
    secret = tmp_path / "secret.txt"
    secret.write_text("TOPSECRET", encoding="utf-8")
    payload = (
        '<?xml version="1.0"?>'
        f'<!DOCTYPE CFXML [<!ENTITY xxe SYSTEM "file:///{secret.as_posix()}">]>'
        "<CFXML><CFData><ProductLineItem>"
        "<ProductLineNumber>&xxe;</ProductLineNumber>"
        "<TransactionType>NEW</TransactionType>"
        "</ProductLineItem></CFData></CFXML>"
    ).encode("utf-8")
    try:
        quote = xml_reader.parse_bytes(payload)
    except xml_reader.QuotationXmlError:
        return  # 파싱 거부도 정상
    assert "TOPSECRET" not in str(quote.groups[0].items[0].line_number)


def test_parse_bytes_needs_bytes_not_text(fixtures):
    """str 로 넘기면 EUC-KR 선언을 존중할 수 없다. 바이트만 받는다."""
    text = (fixtures / "euckr_quote.xml").read_bytes().decode("euc-kr")
    with pytest.raises((ValueError, TypeError)):
        xml_reader.parse_bytes(text)  # type: ignore[arg-type]


# --- 생성 --------------------------------------------------------------------

@pytest.mark.parametrize("name", FIXTURE_NAMES)
def test_bytes_output_matches_path_output(name, tmp_path, fixtures, template_bytes,
                                          template_path):
    """두 경로의 산출물을 셀 단위로 대조한다 (tools/compare.py 와 같은 기준)."""
    source = fixtures / name

    xml = tmp_path / name
    xml.write_bytes(source.read_bytes())
    from_path = convert.convert(xml, template=template_path, today=TODAY).output

    result = convert.convert_bytes(source.read_bytes(), template_bytes,
                                   today=TODAY, source_name=name)
    from_bytes = tmp_path / "bytes.xlsx"
    from_bytes.write_bytes(result.xlsx)

    report = cmp_mod.compare(from_path, from_bytes, set())
    assert report.ok, "\n".join(str(d) for d in report.diffs[:20])


def test_convert_bytes_reports_names_and_counts(fixtures, template_bytes):
    result = convert.convert_bytes(
        (fixtures / "new_quote.xml").read_bytes(), template_bytes,
        today=TODAY, source_name="견적 (수정본)#2.xml")
    assert result.filename == "견적 (수정본)#2.xlsx"
    assert result.group_count == 2
    assert result.line_count == 3
    assert result.elapsed_ms >= 0


def test_convert_bytes_keeps_total_sheet_drawings(fixtures, template_bytes):
    """openpyxl 은 저장 시 도형을 버린다. TOTAL 시트의 로고·머리글이 남아야 한다.

    상세 시트에는 골든과 마찬가지로 그림이 없어야 한다.
    """
    result = convert.convert_bytes((fixtures / "new_quote.xml").read_bytes(),
                                   template_bytes, today=TODAY)
    with zipfile.ZipFile(BytesIO(result.xlsx)) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
    drawn = _sheets_with_drawings(result.xlsx)

    assert media, "템플릿의 로고 이미지가 옮겨지지 않았다"
    assert "xl/worksheets/sheet1.xml" in drawn, (
        f"TOTAL(sheet1)에 그림이 없다: {drawn}")
    assert _shape_count(result.xlsx) == _shape_count(template_bytes), (
        "머리글 도형이 통째로 옮겨지지 않았다 (openpyxl 은 도형을 만들지 못한다)")

    # 상세 시트(sheet2·sheet3)는 어느 환경에서도 비어 있어야 한다.
    details = {"xl/worksheets/sheet2.xml", "xl/worksheets/sheet3.xml"}
    assert not details & set(drawn), f"상세 시트에 그림이 남았다: {drawn}"


@pytest.mark.skipif(
    HAS_PILLOW,
    reason="Pillow 가 있으면 openpyxl 이 숨김 template 시트의 로고도 다시 쓴다. "
           "배포되는 세 갈래에는 Pillow 가 없다",
)
def test_convert_bytes_draws_only_on_the_total_sheet(fixtures, template_bytes):
    """운영 환경(Pillow 없음)에서는 그림이 TOTAL 한 장에만 남는다."""
    result = convert.convert_bytes((fixtures / "new_quote.xml").read_bytes(),
                                   template_bytes, today=TODAY)
    drawn = _sheets_with_drawings(result.xlsx)
    assert drawn == ["xl/worksheets/sheet1.xml"], (
        f"그림은 TOTAL(sheet1)에만 있어야 한다: {drawn}")


def test_carry_over_bytes_is_a_noop_without_drawings(fixtures, template_bytes):
    """그림이 없는 템플릿이면 손대지 않고 None 을 돌려준다."""
    plain_template = _without_drawings(template_bytes)
    assert _sheets_with_drawings(plain_template) == [], "입력부터 그림이 없어야 한다"

    quote = xml_reader.parse_bytes((fixtures / "new_quote.xml").read_bytes())
    built = BytesIO()
    ibm_writer.build(quote, BytesIO(plain_template), today=TODAY).save(built)
    assert drawings.carry_over_bytes(plain_template, built.getvalue()) is None


def test_build_bytes_produces_a_readable_workbook(fixtures, template_bytes):
    from openpyxl import load_workbook

    quote = xml_reader.parse_bytes((fixtures / "upgrade_quote.xml").read_bytes())
    data = ibm_writer.build_bytes(quote, template_bytes, today=TODAY)
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["TOTAL", "SERVER 1", "template"]
    assert wb["TOTAL"]["C3"].value == TODAY.isoformat()

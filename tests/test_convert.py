"""변환 오케스트레이션 (경로 입력) — 데스크톱 경로와 공용 코어의 계약."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest
from openpyxl import load_workbook

from quotation.core import convert, resources, xml_reader
from quotation.core.xml_reader import QuotationXmlError

TODAY = dt.date(2026, 7, 23)


def _convert(tmp_path: Path, source: Path, **kwargs) -> convert.Result:
    """fixture 를 tmp_path 로 복사해 변환한다. 산출물은 그 옆에 생긴다."""
    xml = tmp_path / source.name
    xml.write_bytes(source.read_bytes())
    return convert.convert(xml, today=TODAY, **kwargs)


# --- 리소스 ------------------------------------------------------------------

def test_default_template_ships_with_the_core_package():
    assert resources.default_template_path().is_file()
    assert resources.default_template_bytes()[:2] == b"PK", "xlsx(zip) 여야 한다"


def test_there_is_exactly_one_template():
    """양식은 하나뿐이다. 데스크톱과 웹이 같은 파일을 쓴다.

    다른 양식(과거의 삼성 SDS B2B 등)은 지원하지 않는다. 양식이 늘어나면
    두 실행 경로의 산출물이 갈리므로 여기서 막는다.
    """
    books = sorted(p.name for p in resources.RESOURCE_DIR.glob("*.xlsx"))
    assert books == [resources.TEMPLATE_NAME], f"양식이 하나가 아니다: {books}"


# --- 출력 위치와 이름 ----------------------------------------------------------

def test_output_location_is_fixed(tmp_path, fixtures):
    """저장 위치는 언제나 XML 과 같은 폴더다. 바꿀 수단이 없다."""
    nested = tmp_path / "깊은" / "폴더"
    nested.mkdir(parents=True)
    result = _convert(nested, fixtures / "new_quote.xml")
    assert result.output == nested / "new_quote.xlsx"
    assert result.output.is_file()


@pytest.mark.parametrize("name,expected", [
    ("a.xml", "a.xlsx"),
    ("견적 (수정본)#2.XML", "견적 (수정본)#2.xlsx"),
    ("점.이.여럿.xml", "점.이.여럿.xlsx"),
])
def test_output_name_keeps_the_stem(name, expected):
    assert convert.output_name_for(name) == expected


# --- 변환 --------------------------------------------------------------------

def test_convert_reports_groups_and_progress(tmp_path, fixtures):
    seen: list[tuple[int, str]] = []
    result = _convert(tmp_path, fixtures / "new_quote.xml",
                      progress=lambda p, m: seen.append((p, m)))
    assert result.group_count == 2
    assert seen[0][0] < seen[-1][0] == 100
    # 상태 메시지는 원본 프로그램의 문구를 유지한다
    assert seen[-1][1] == "견적서작성을 완료하였습니다."


def test_sheets_are_total_details_and_hidden_template(tmp_path, fixtures):
    wb = load_workbook(_convert(tmp_path, fixtures / "new_quote.xml").output)
    assert wb.sheetnames == ["TOTAL", "SAMPLE-100 #1", "SAMPLE-200 #2", "template"]
    assert wb["template"].sheet_state == "hidden"


def test_total_sheet_amounts(tmp_path, fixtures):
    """H/W 구간 2,601 (1,000.50x2 + 100x3x2) / S/W 구간 250 / 두 번째 장비군 11,800."""
    wb = load_workbook(_convert(tmp_path, fixtures / "new_quote.xml").output)
    ws = wb["TOTAL"]
    assert ws["G8"].value == 2601
    assert ws["G9"].value == 250
    assert ws["G10"].value == "=SUM(G8:G9)"
    assert ws["G11"].value == 11800
    assert ws["C3"].value == TODAY.isoformat()


def test_no_charge_is_written_as_text(tmp_path, fixtures):
    """무상 품목은 금액 셀에 "N/C" 문자열로 남고 합계에서는 0 이다."""
    wb = load_workbook(_convert(tmp_path, fixtures / "no_charge.xml").output)
    ws = wb["SAMPLE-300 #1"]
    assert ws["F8"].value == "N/C" and ws["G8"].value == "N/C"
    # 기준행에 가격이 없으면 서브 수량은 부모 참조 없는 수식이다
    assert ws["E9"].value == "=4"
    assert wb["TOTAL"]["G8"].value is None, "합계가 0 이면 병합 셀을 비운다"


def test_removed_parts_are_negative_and_red(tmp_path, fixtures):
    from quotation.core.writer.decorate import RED

    wb = load_workbook(_convert(tmp_path, fixtures / "upgrade_quote.xml").output)
    ws = wb["SERVER 1"]
    assert wb.sheetnames == ["TOTAL", "SERVER 1", "template"], "증설은 한 장이다"

    # 8행 UPGRADE 본체 / 9행 ADD / 10행 REMOVE
    assert ws["E10"].value == "=-1*E8"
    assert ws["F10"].value is None and ws["G10"].value is None
    for col in "CDEFG":
        color = ws[f"{col}10"].font.color
        assert color is not None and color.rgb == RED, f"{col}10 이 붉지 않다"
    assert ws["C8"].font.color is None, "정상 행이 물들었다"


def test_euckr_document_is_accepted(tmp_path, fixtures):
    """2005년 형식(EUC-KR, 인라인 DTD)도 계속 읽혀야 한다."""
    wb = load_workbook(_convert(tmp_path, fixtures / "euckr_quote.xml").output)
    assert "서버 1" in wb.sheetnames
    assert wb["TOTAL"]["G8"].value == 1000.5


def test_conversion_preserves_component_quantity_and_lp(tmp_path, fixtures):
    """서브라인 수량은 부모 수량 참조 수식, 단가는 원본 소수 자릿수 그대로."""
    wb = load_workbook(_convert(tmp_path, fixtures / "new_quote.xml").output)
    ws = wb["SAMPLE-100 #1"]
    assert [ws[f"{col}8"].value for col in "CDEFG"] == [
        "1000-A1B", "SAMPLE-100 #1:Sample Storage Control Enclosure", 2,
        1000.5, "=E8*F8",
    ]
    assert [ws[f"{col}9"].value for col in "CDE"] == [
        "AC01", "Sample Cache Module", "=3*E8",
    ]


def test_maintenance_is_never_written(tmp_path, fixtures):
    """유지정비료를 견적서 H·I열에 쓰지 않는다."""
    wb = load_workbook(_convert(tmp_path, fixtures / "new_quote.xml").output)
    for sheet in wb.worksheets:
        for row in range(8, 40):
            for col in ("H", "I"):
                value = sheet[f"{col}{row}"].value
                assert value is None, (
                    f"[{sheet.title}] {col}{row} 에 유지정비료가 남아 있다: {value!r}")


def test_supply_row_is_always_blank(tmp_path, fixtures):
    """공급가 행은 언제나 공란이다. 할인 기능은 없앴다."""
    wb = load_workbook(_convert(tmp_path, fixtures / "new_quote.xml").output)
    ws = wb["TOTAL"]
    grand = next(r for r in range(8, 40)
                 if ws[f"B{r}"].value == "총        합       계")
    assert str(ws[f"G{grand}"].value).startswith("=SUM(")
    assert ws[f"B{grand + 1}"].value == "공        급       가"
    assert ws[f"G{grand + 1}"].value is None


# --- 견적서 번호 ---------------------------------------------------------------

def _quote_number(tmp_path, fixtures, template_b2: str) -> str:
    from quotation.core.writer import ibm_writer

    template = tmp_path / "t.xlsx"
    wb = load_workbook(resources.default_template_path())
    wb["TOTAL"]["B2"] = template_b2
    wb.save(template)

    quote = xml_reader.parse(fixtures / "new_quote.xml")
    out = ibm_writer.write(quote, template, tmp_path / "o.xlsx", today=TODAY)
    return load_workbook(out)["TOTAL"]["B2"].value


def test_quote_number_updates_year_only(tmp_path, fixtures):
    """결과는 언제나 'NO : Trialinfo-{YY}-' 다. 연도 뒤에 아무것도 붙지 않는다."""
    assert _quote_number(tmp_path, fixtures,
                         "NO : Trialinfo-24-") == "NO : Trialinfo-26-"


def test_quote_number_keeps_custom_prefix(tmp_path, fixtures):
    """앞부분 문구는 템플릿에 적힌 그대로 둔다."""
    assert _quote_number(tmp_path, fixtures,
                         "견적 NO : Trialinfo-20-") == "견적 NO : Trialinfo-26-"


# --- 오류 --------------------------------------------------------------------

def test_bad_xml_reports_original_message(tmp_path):
    bad = tmp_path / "bad.xml"
    bad.write_text("<CFXML><CFData/></CFXML>", encoding="utf-8")
    with pytest.raises(QuotationXmlError, match="Item을 찾을 수 없습니다"):
        convert.convert(bad, today=TODAY)

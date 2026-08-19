"""통합 모드 (레노버 x86) 검증.

기대값의 근거는 `quotation/core/integrated.py` 머리말에 적어 둔 Lenovo DCSC
요약표 대조다. fixture 는 `tests/fixtures/public/integrated_quote.xml` 로,
실데이터가 아니라 그 구조만 남긴 익명화 자료다.

여기서 지키려는 것은 셋이다.
  1. UNIX 모드는 한 줄도 달라지지 않는다
  2. 같은 기종 여러 대의 시트명이 겹치지 않는다
  3. 본체 LP 에 이미 들어 있는 소프트웨어·서비스 금액을 두 번 세지 않는다
"""
from __future__ import annotations

import datetime as dt
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from quotation.core import modes, xml_reader
from quotation.core.naming import (product_key, safe_sheet_name, sheet_name,
                                   unique_sheet_names)
from quotation.core.writer import ibm_writer

TODAY = dt.date(2026, 8, 19)


@pytest.fixture(scope="module")
def source(fixtures) -> bytes:
    return (fixtures / "integrated_quote.xml").read_bytes()


@pytest.fixture(scope="module")
def unix(source):
    return xml_reader.parse_bytes(source, mode=modes.UNIX)


@pytest.fixture(scope="module")
def integrated(source):
    return xml_reader.parse_bytes(source, mode=modes.INTEGRATED)


# --- 모드 값 ------------------------------------------------------------------

def test_default_mode_is_unix():
    assert modes.DEFAULT == modes.UNIX
    assert modes.normalize(None) == modes.UNIX
    assert modes.normalize("") == modes.UNIX
    assert modes.normalize(" INTEGRATED ") == modes.INTEGRATED


def test_unknown_mode_is_rejected():
    with pytest.raises(ValueError):
        modes.normalize("lenovo")


def test_parse_defaults_to_unix(source, unix):
    assert xml_reader.parse_bytes(source).groups == unix.groups


# --- 시트명 -------------------------------------------------------------------

def test_same_model_twice_collides_in_unix_mode(unix):
    """지금까지의 동작. 이름이 겹치는 것을 openpyxl 이 뒤늦게 얼버무린다."""
    names = [g.sheet_name for g in unix.groups]
    assert names[0] == names[1] == "SAMPLESYSTEM SX100 V4-3YR BASE"


def test_integrated_names_sheets_after_the_product_name(integrated):
    assert [g.sheet_name for g in integrated.groups] == [
        "백업서버_1식", "메일-스펨_2식", "SAMPLE 42U DEEP STATIC RACK"]
    assert len({g.sheet_name for g in integrated.groups}) == 3


def test_integrated_keeps_the_written_name_for_the_item_column(integrated):
    """종목 칸과 상세 시트 제목에는 사람이 적은 이름을 그대로 쓴다."""
    assert integrated.groups[1].item_key == "메일/스펨_2식"
    assert integrated.groups[1].title == "메일/스펨_2식"
    # 시트명만 Excel 금칙 문자를 바꾼다
    assert integrated.groups[1].sheet_name == "메일-스펨_2식"


def test_group_without_a_product_name_falls_back_to_the_description(integrated):
    rack = integrated.groups[2]
    assert rack.item_key == "Sample 42U Deep Static Rack"
    assert rack.sheet_name == "SAMPLE 42U DEEP STATIC RACK"


@pytest.mark.parametrize("raw,expected", [
    ("메일/스펨_1식", "메일-스펨_1식"),
    ("a:b\\c/d?e*f[g]h", "A-B-C-D-E-F-G-H"),
    ("   ", "SHEET"),
    ("history", "SHEET"),
    ("'따옴표'", "따옴표"),
])
def test_safe_sheet_name(raw, expected):
    assert safe_sheet_name(raw) == expected


def test_safe_sheet_name_never_exceeds_the_excel_limit():
    assert len(safe_sheet_name("가" * 60)) == 31


def test_unique_sheet_names_splits_repeats():
    assert unique_sheet_names(["백업서버", "백업서버", "백업서버", "웹서버"]) == [
        "백업서버", "백업서버 (2)", "백업서버 (3)", "웹서버"]


def test_unique_sheet_names_keeps_the_31_character_limit():
    long = "가" * 31
    made = unique_sheet_names([long, long])
    assert made[0] == long
    assert made[1].endswith(" (2)") and len(made[1]) == 31


def test_product_key_prefers_the_product_name():
    assert product_key("백업서버_1식", "SampleSystem SX100") == "백업서버_1식"
    assert product_key("", "SampleSystem SX100") == "SampleSystem SX100"
    assert product_key("   ", "IBM Power E1080") == "Power E1080"


def test_unix_sheet_name_is_unchanged():
    assert sheet_name("Server 1") == "SERVER 1"


# --- 금액 ---------------------------------------------------------------------

def test_unix_mode_adds_software_and_services_on_top(unix):
    """지금까지의 동작. 레노버 문서에서는 이것이 이중 계상이 된다."""
    assert unix.groups[0].amount() == Decimal("45863001")


def test_integrated_counts_the_body_line_only(integrated):
    """본체 LP 하나가 그 서버의 전체 금액이다 (integrated.py 머리말)."""
    assert integrated.groups[0].amount() == Decimal("40172000")
    # 수량 2대는 본체 LP 의 두 배다. 서비스는 이미 그 안에 있다.
    assert integrated.groups[1].amount() == Decimal("85624000")


def test_integrated_empties_only_the_price_not_the_line(integrated):
    """무엇이 들어 있는지는 남기고 금액 칸만 비운다."""
    group = integrated.groups[0]
    assert [i.line_number for i in group.items] == ["1000", "2000", "3000"]
    assert group.items[0].unit_price == Decimal("4.0172E+7")
    assert group.items[1].unit_price is None   # 자리표 1 원
    assert group.items[2].unit_price is None   # 본체 LP 에 이미 들어 있다


def test_integrated_leaves_parts_only_groups_alone(integrated):
    """랙·PDU 처럼 본체 없이 부품만 있는 그룹은 제 금액을 그대로 갖는다."""
    assert integrated.groups[2].amount() == Decimal("17920000")


# --- 워크북 -------------------------------------------------------------------

@pytest.fixture(scope="module")
def workbook(integrated, template_bytes):
    xlsx = ibm_writer.build_bytes(integrated, template_bytes, today=TODAY)
    return load_workbook(BytesIO(xlsx))


def test_workbook_has_one_sheet_per_group_with_distinct_names(workbook):
    assert workbook.sheetnames == [
        "TOTAL", "백업서버_1식", "메일-스펨_2식", "SAMPLE 42U DEEP STATIC RACK",
        "template"]
    assert all(len(name) <= 31 for name in workbook.sheetnames)


def test_total_sheet_prices_the_body_line_only(workbook):
    total = workbook["TOTAL"]
    assert total["B8"].value == "백업서버_1식"
    assert total["G8"].value == 40172000        # H/W 구간 = 본체 LP
    assert total["G9"].value is None            # S/W 구간은 금액을 비운다
    assert total["C9"].value == "7SXX-CTOBWW"   # 라인 자체는 남는다
    assert total["C10"].value == "5WS7C00001"


def test_detail_sheet_title_keeps_the_forbidden_character(workbook):
    assert workbook["메일-스펨_2식"]["C1"].value == "(메일/스펨_2식)"


def test_detail_sheet_software_total_is_zero(workbook):
    """S/W 합계는 0 이다. 금액은 H/W 본체 한 줄에만 있다."""
    detail = workbook["백업서버_1식"]
    labels = {c.value: c.row for c in detail["C"] if isinstance(c.value, str)}
    software_total = labels[ibm_writer.LBL_SW_TOTAL]
    hardware_total = labels[ibm_writer.LBL_HW_TOTAL]
    assert detail[f"G{hardware_total}"].value.startswith("=SUM(")
    assert detail[f"G{software_total}"].value.startswith("=SUM(")
    # S/W 구간의 단위가 칸은 비어 있다
    for row in range(software_total - 2, software_total):
        assert detail[f"F{row}"].value is None

"""표 장식 — 테두리·채우기·정렬·행높이.

원본 VB6 의 `DecorateTOTSheet` 에 해당한다. EXE 문자열에 남아 있던 범위들
(`B8:H`, `B8:B`, `C8:C`, `F8:F`, `G8:G`, `H8:I`, `E8:E`)이 이 작업의 흔적이다.

병합은 장식이 끝난 뒤에 적용해야 한다. openpyxl 은 병합 시 덮인 셀의 스타일을
버리는데, 골든도 정확히 그 상태다.
"""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

#: 색인 64 = 자동(automatic). 골든의 모든 테두리가 이 색을 명시하고 있다.
AUTO = Color(indexed=64)
THIN = Side(style="thin", color=AUTO)
MEDIUM = Side(style="medium", color=AUTO)
#: 헤더 아래 굵은 선만 색인 0(검정)을 쓴다. 골든이 그렇게 저장되어 있다.
MEDIUM_BLACK = Side(style="medium", color=Color(indexed=0))

#: 레거시 팔레트 색인 9 = 흰색. 골든이 이 형태로 저장되어 있다.
FILL_BODY = PatternFill("solid", fgColor=Color(indexed=9), bgColor=Color(indexed=64))
FILL_YELLOW = PatternFill("solid", fgColor="FFFFFF99", bgColor=Color(indexed=64))
FILL_CYAN = PatternFill("solid", fgColor="FFCCFFFF", bgColor=Color(indexed=64))
FILL_ORANGE = PatternFill("solid", fgColor="FFFFCC99", bgColor=Color(indexed=64))

#: Software 구간은 파란 글꼴이다.
BLUE = "FF0000FF"
#: 제거(REMOVE) 부품은 붉은 글꼴이다.
RED = "FFFF0000"

ROW_H_DATA = 12.0
ROW_H_TOTAL = 20.1
ROW_H_TRAILER = 30.0

COLS = "BCDEFGH"

#: 열별 가로 정렬. 세로는 전부 가운데다. 줄바꿈(wrap)은 템플릿 값을 그대로 둔다
#: — 골든의 wrap 은 셀마다 제각각이라 규칙이 아니라 템플릿에서 물려받은 것이다.
HALIGN = {"C": "center", "D": "left", "E": "center",
          "F": "right", "G": "right", "H": "right"}


@dataclass
class Layout:
    """장식에 필요한 표 구조. writer 가 값을 쓰면서 채운다."""

    first_row: int = 8
    #: B열 세로 병합 구간 (시작, 끝)
    bands: list[tuple[int, int]] = field(default_factory=list)
    #: C:F 로 병합되는 행 (합계류)
    cf_rows: list[int] = field(default_factory=list)
    #: B:F 로 병합되는 행 (총합계, 공급가)
    bf_rows: list[int] = field(default_factory=list)
    #: 노랑 채우기 (C, G, H)
    yellow_rows: list[int] = field(default_factory=list)
    #: 하늘 채우기 (C, G, H)
    cyan_rows: list[int] = field(default_factory=list)
    #: B열 노랑 채우기 셀의 행
    yellow_labels: list[int] = field(default_factory=list)
    #: 파란 글꼴 행
    blue_rows: list[int] = field(default_factory=list)
    #: 파란 글꼴을 B열(구간 라벨)까지 적용하는가. 상세 시트만 True.
    blue_includes_label: bool = False
    #: 제거(REMOVE) 부품 행. 붉은 글꼴로 표시한다.
    red_rows: list[int] = field(default_factory=list)
    grand_row: int = 0
    supply_row: int = 0
    #: 상세 시트는 헤더 아래 테두리를 medium 으로 다시 그린다 (템플릿은 double)
    fix_header_bottom: bool = False
    #: 표 첫 행 윗선의 색. TOTAL 은 색인 64, 상세 시트는 색인 0 이다.
    top_black: bool = False
    #: 서브라인이 없는 블록의 스페이서 행. C열에 배경을 칠하지 않는다.
    spacer_rows: list[int] = field(default_factory=list)
    #: I열에 배경을 칠할 행 (MA 기간을 적은 행)
    i_filled_rows: list[int] = field(default_factory=list)


def _side(col: str, layout: Layout, row: int) -> tuple[Side | None, Side | None]:
    """가로 방향 테두리 (왼쪽, 오른쪽)."""
    if row in layout.bf_rows:
        # B:F 병합 — 첫 칸은 양쪽, 중간은 없음, 마지막 칸은 오른쪽만
        if col == "B":
            return MEDIUM, THIN
        if col in "CDE":
            return None, None
        if col == "F":
            return None, THIN
    elif row in layout.cf_rows:
        if col == "C":
            return THIN, THIN
        if col in "DE":
            return None, None
        if col == "F":
            return None, THIN
    if col == "B":
        return MEDIUM, THIN
    if col == "H":
        return THIN, MEDIUM
    return THIN, THIN


def _vert(col: str, layout: Layout, row: int) -> tuple[Side | None, Side | None]:
    """세로 방향 테두리 (위, 아래)."""
    first_top = MEDIUM_BLACK if layout.top_black else MEDIUM
    top = first_top if row == layout.first_row else THIN
    bottom = MEDIUM if row == layout.supply_row else THIN

    if col == "B":
        for start, end in layout.bands:
            if start <= row <= end:
                if row == start:
                    return (first_top if start == layout.first_row
                            else THIN), THIN
                if row == end:
                    return None, THIN
                return None, None
    return top, bottom


def decorate(ws: Worksheet, layout: Layout) -> None:
    """표 전체에 테두리·채우기·정렬·행높이를 적용한다."""
    last = layout.supply_row

    for row in range(layout.first_row, last + 1):
        ws.row_dimensions[row].height = (
            ROW_H_TOTAL if row in (layout.grand_row, layout.supply_row)
            else ROW_H_DATA
        )
        for col in COLS:
            cell = ws[f"{col}{row}"]
            left, right = _side(col, layout, row)
            top, bottom = _vert(col, layout, row)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

            # 스페이서 행의 C열은 원본이 손대지 않아 배경이 없다
            if col != "B" and not (col == "C" and row in layout.spacer_rows):
                cell.fill = FILL_BODY
                # 총합계·공급가 행은 가로 정렬을 지정하지 않는다 (골든 = 일반)
                horizontal = None if row in layout.bf_rows else HALIGN[col]
                cell.alignment = Alignment(
                    horizontal=horizontal, vertical="center",
                    wrap_text=cell.alignment.wrap_text)

    # 데이터 첫 행과 MA 기간을 적은 행은 I열까지 배경이 이어진다
    for row in {layout.first_row, *layout.i_filled_rows}:
        cell = ws[f"I{row}"]
        cell.fill = FILL_BODY
        cell.alignment = Alignment(vertical="center")

    _apply_row_fills(ws, layout)
    _apply_blue(ws, layout)
    _decorate_trailer(ws, layout)
    if layout.fix_header_bottom:
        _fix_header(ws)


def _apply_row_fills(ws: Worksheet, layout: Layout) -> None:
    for row in layout.yellow_rows:
        for col in "CGH":
            ws[f"{col}{row}"].fill = FILL_YELLOW
    for row in layout.cyan_rows:
        for col in "CGH":
            ws[f"{col}{row}"].fill = FILL_CYAN
    for row in layout.yellow_labels:
        ws[f"B{row}"].fill = FILL_YELLOW
    for row in (layout.grand_row, layout.supply_row):
        if row:
            for col in "BGH":
                ws[f"{col}{row}"].fill = FILL_ORANGE
            for col in "CDEF":
                ws[f"{col}{row}"].fill = PatternFill()


def _recolor(ws: Worksheet, rows, cols: str, color: str) -> None:
    for row in rows:
        for col in cols:
            cell = ws[f"{col}{row}"]
            f = cell.font
            if f.name is None:
                continue
            cell.font = Font(name=f.name, size=f.size, bold=f.bold, color=color)


def _apply_blue(ws: Worksheet, layout: Layout) -> None:
    """Software 구간은 파랑, 제거 부품은 빨강.

    파랑은 TOTAL 시트에서 C~G 만, 상세 시트에서는 구간 라벨(B)까지 적용한다.
    빨강은 파랑보다 나중에 칠한다. S/W 구간에서 제거된 부품도 빨갛게 나와야 한다.
    """
    _recolor(ws, layout.blue_rows,
             "BCDEFG" if layout.blue_includes_label else "CDEFG", BLUE)
    _recolor(ws, layout.red_rows, "CDEFG", RED)


def _decorate_trailer(ws: Worksheet, layout: Layout) -> None:
    """공급가 아래 비고 블록 (2행). 바깥만 medium 이다."""
    top = layout.supply_row + 1
    bottom = top + 1
    ws.row_dimensions[top].height = ROW_H_TRAILER
    ws.row_dimensions[bottom].height = ROW_H_TRAILER

    ws[f"B{top}"].border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM,
                                  bottom=MEDIUM)
    ws[f"B{top}"].fill = FILL_BODY
    ws[f"B{bottom}"].border = Border(left=MEDIUM, bottom=MEDIUM)
    for col in "CDEFG":
        ws[f"{col}{top}"].border = Border(top=MEDIUM)
        ws[f"{col}{bottom}"].border = Border(bottom=MEDIUM)
    ws[f"H{top}"].border = Border(right=MEDIUM, top=MEDIUM)
    ws[f"H{bottom}"].border = Border(right=MEDIUM, bottom=MEDIUM)


def _fix_header(ws: Worksheet) -> None:
    """상세 시트 헤더 아래를 medium 으로. 템플릿은 double 로 되어 있다."""
    for addr in [f"{c}7" for c in COLS] + ["D6", "H6"]:
        b = ws[addr].border
        ws[addr].border = Border(left=b.left, right=b.right, top=b.top,
                                 bottom=MEDIUM_BLACK)

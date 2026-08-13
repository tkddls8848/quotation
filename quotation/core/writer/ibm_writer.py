"""IBM 견적서 워크북 생성."""
from __future__ import annotations

import datetime as dt
from copy import copy
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import IO

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Group, LineItem, Quotation, SubLineItem
from ..money import NO_CHARGE, Amount, is_priced
from . import drawings
from .decorate import Layout, decorate

SHEET_TOTAL = "TOTAL"
SHEET_TEMPLATE = "template"

FIRST_DATA_ROW = 8
TRAILER_ROWS = 2

# --- 라벨 (공백 개수까지 원본과 동일해야 한다) ----------------------------------
LBL_SUBTOTAL = "합                   계"
LBL_HW_TOTAL = "합                   계(HardWare)"
LBL_SW_TOTAL = "합                   계(SoftWare)"
LBL_GRAND = "총        합       계"
LBL_SUPPLY = "공        급       가"
LBL_HW = "H/W"
LBL_SW = "S/W"

# --- 서식 --------------------------------------------------------------------
FMT_TEXT = "@"
FMT_TOTAL_NUM = "#,##0_);[Red]\\(#,##0\\)"
FMT_DETAIL_NUM = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
FONT_DATA = Font(name="Tahoma", size=9)
FONT_DATA_BOLD = Font(name="Tahoma", size=9, bold=True)
FONT_LABEL = Font(name="돋움", size=9, bold=True)
# 원본 EXE 문자열의 "Tahoma" / "HY헤드라인M" 이 이 두 셀을 가리킨다.
# 템플릿은 정반대(C1=HY헤드라인M, C3=Tahoma/10)이므로 반드시 덮어써야 한다.
FONT_TITLE = Font(name="Tahoma", size=18, bold=True)
FONT_DATE = Font(name="HY헤드라인M", size=9)

LEFT = Alignment(horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
GENERAL = Alignment(vertical="center")


def _put(ws: Worksheet, coord: str, value, *, fmt: str | None = None,
         align: Alignment | None = None, font: Font | None = None,
         bold: bool | None = None):
    """셀 기록. font 를 주지 않으면 템플릿 글꼴을 유지하고 bold 만 조정한다."""
    cell = ws[coord]
    cell.value = value
    if fmt is not None:
        cell.number_format = fmt
    if align is not None:
        cell.alignment = align
    if font is not None:
        cell.font = font
    elif bold is not None:
        kept = copy(cell.font)
        kept.b = bold
        cell.font = kept
    return cell


def _num(value: Decimal):
    """Decimal -> Excel 수치. 정수는 int 로 써야 골든과 일치한다."""
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _amount_cell(ws: Worksheet, coord: str, amount: Amount, fmt: str):
    if amount is NO_CHARGE:
        _put(ws, coord, "N/C", fmt=fmt, font=FONT_DATA)
    elif is_priced(amount):
        _put(ws, coord, _num(amount), fmt=fmt, font=FONT_DATA)


def _update_quote_number(ws: Worksheet, today: dt.date) -> None:
    """B2의 `Trialinfo-YY-` 번호에서 연도만 갱신한다."""
    prefix = ws["B2"].value.split("Trialinfo-", 1)[0]
    ws["B2"] = f"{prefix}Trialinfo-{today:%y}-"


def _merge(pending: list[str], col: str, top: int, bottom: int):
    if bottom > top:
        pending.append(f"{col}{top}:{col}{bottom}")


def _merge_across(pending: list[str], left: str, right: str, row: int):
    pending.append(f"{left}{row}:{right}{row}")


# --- TOTAL 시트 ---------------------------------------------------------------

def _write_total_sheet(ws: Worksheet, quote: Quotation, today: dt.date):
    merges: list[str] = []
    lay = Layout(first_row=FIRST_DATA_ROW)

    _update_quote_number(ws, today)
    _put(ws, "C3", today.isoformat(), fmt=FMT_TEXT, align=LEFT, font=FONT_DATE)

    row = FIRST_DATA_ROW
    subtotal_rows: list[int] = []

    for group in quote.groups:
        group_start = row

        for kind, items in group.sections():
            sec_start = row
            for item in items:
                _put(ws, f"C{row}", item.part_number, fmt=FMT_TEXT, font=FONT_DATA)
                _put(ws, f"D{row}", item.description, font=FONT_DATA)
                _put(ws, f"E{row}", item.quantity, font=FONT_DATA)
                if kind != "Hardware":
                    lay.blue_rows.append(row)
                row += 1
            sec_end = row - 1

            amount = sum((i.amount() for i in items), Decimal(0))
            if amount != 0:
                _put(ws, f"F{sec_start}", f"=G{sec_start}/E{sec_start}",
                     fmt=FMT_TOTAL_NUM, font=FONT_DATA)
                _put(ws, f"G{sec_start}", _num(amount),
                     fmt=FMT_TOTAL_NUM, font=FONT_DATA)
            _merge(merges, "F", sec_start, sec_end)
            _merge(merges, "G", sec_start, sec_end)

        group_end = row - 1

        _merge(merges, "H", group_start, group_end)

        _put(ws, f"C{row}", LBL_SUBTOTAL, fmt=FMT_TEXT, font=FONT_LABEL)
        _merge_across(merges, "C", "F", row)
        _put(ws, f"G{row}", f"=SUM(G{group_start}:G{group_end})",
             fmt=FMT_TOTAL_NUM, font=FONT_DATA_BOLD)
        subtotal_rows.append(row)
        lay.cf_rows.append(row)
        lay.cyan_rows.append(row)

        _put(ws, f"B{group_start}", group.item_key, align=CENTER_WRAP,
             font=FONT_DATA_BOLD)
        _merge(merges, "B", group_start, row)
        lay.bands.append((group_start, row))
        row += 1

    _write_footer(ws, row, subtotal_rows, FMT_TOTAL_NUM, merges, lay,
                  always_sum=True)
    return merges, lay


# --- 상세 시트 ----------------------------------------------------------------

def _write_detail_sheet(ws: Worksheet, group: Group, today: dt.date):
    merges: list[str] = []
    lay = Layout(first_row=FIRST_DATA_ROW, fix_header_bottom=True,
                 blue_includes_label=True, top_black=True)

    _put(ws, "C1", f"({group.sheet_name})", fmt=FMT_TEXT, align=CENTER_WRAP,
         font=FONT_TITLE)
    _put(ws, "C3", today.isoformat(), fmt=FMT_TEXT, align=LEFT, font=FONT_DATE)

    row = FIRST_DATA_ROW
    section_total_rows: list[int] = []

    for kind, items in group.sections():
        sec_start = row
        is_hw = kind == "Hardware"
        # 블록별 합계행은 H/W 구간에 라인이 둘 이상일 때만 쓴다.
        # 하나뿐이면 바로 합계(HardWare) 로 가고 수식도 범위형이다.
        # S/W 는 개수와 무관하게 언제나 범위형 하나다.
        per_block = is_hw and len(items) > 1
        block_rows: list[int] = []

        for item in items:
            block_start = row
            row = _write_item_block(ws, item, row, is_hw, lay)
            if per_block:
                # 서브라인이 없으면 스페이서 1행을 둔다
                last = max(row - 1, block_start + 1)
                if last >= row:
                    lay.spacer_rows.extend(range(row, last + 1))
                row = last + 1
                _put(ws, f"C{row}", LBL_SUBTOTAL, fmt=FMT_TEXT, font=FONT_LABEL)
                _merge_across(merges, "C", "F", row)
                _put(ws, f"G{row}", f"=SUM(G{block_start}:G{last})",
                     fmt=FMT_DETAIL_NUM, font=FONT_DATA)
                block_rows.append(row)
                lay.cf_rows.append(row)
                lay.yellow_rows.append(row)
                row += 1

        _put(ws, f"C{row}", LBL_HW_TOTAL if is_hw else LBL_SW_TOTAL,
             fmt=FMT_TEXT, font=FONT_LABEL)
        _merge_across(merges, "C", "F", row)
        if per_block:
            _put(ws, f"G{row}", f"=SUM({','.join(f'G{r}' for r in block_rows)})",
                 fmt=FMT_DETAIL_NUM, font=FONT_DATA)
        else:
            _put(ws, f"G{row}", f"=SUM(G{sec_start}:G{row - 1})",
                 fmt=FMT_DETAIL_NUM, font=FONT_DATA)
        section_total_rows.append(row)
        lay.cf_rows.append(row)
        lay.cyan_rows.append(row)

        _put(ws, f"B{sec_start}", LBL_HW if is_hw else LBL_SW, align=CENTER,
             font=FONT_DATA_BOLD)
        _merge(merges, "B", sec_start, row)
        lay.bands.append((sec_start, row))
        lay.yellow_labels.append(sec_start)
        if not is_hw:
            lay.blue_rows.extend(range(sec_start, row + 1))
        row += 1

    _write_footer(ws, row, section_total_rows, FMT_DETAIL_NUM, merges, lay)
    return merges, lay


def _write_item_block(ws: Worksheet, item: LineItem, row: int, is_hw: bool,
                      lay: Layout) -> int:
    """ProductLineItem 1건과 그 서브라인을 기록한다."""
    base = row
    _put(ws, f"C{row}", item.part_number, fmt=FMT_TEXT, font=FONT_DATA)
    _put(ws, f"D{row}", item.description, font=FONT_DATA)
    _put(ws, f"E{row}", item.quantity, font=FONT_DATA)
    _amount_cell(ws, f"F{row}", item.unit_price, FMT_DETAIL_NUM)
    _write_g(ws, row, item.unit_price)
    row += 1

    for sub in item.subs:
        _put(ws, f"C{row}", sub.part_number, fmt=FMT_TEXT, font=FONT_DATA)
        _put(ws, f"D{row}", sub.description, font=FONT_DATA)
        _put(ws, f"E{row}", _sub_quantity(sub, base, item, is_hw),
             font=FONT_DATA)
        _amount_cell(ws, f"F{row}", sub.unit_price, FMT_DETAIL_NUM)
        _write_g(ws, row, sub.unit_price)
        if sub.is_removal:
            lay.red_rows.append(row)
        row += 1

    return row


def _sub_quantity(sub: SubLineItem, base: int, item: LineItem, is_hw: bool):
    """서브라인 수량 셀. 골든이 세 형태를 쓴다 (SPEC_CELLMAP.md §4.5).

        H/W + 기준행에 가격 있음 -> "=8*E8"   부모 수량 상대 참조
        H/W + 기준행이 N/C       -> "=1"      참조 없는 수식
        S/W                      -> 1         상수

    제거(REMOVE) 부품은 부호를 뒤집는다. XML 의 수량은 양수로 들어온다.
    """
    quantity = -sub.quantity if sub.is_removal else sub.quantity
    if not is_hw:
        return quantity
    if is_priced(item.unit_price):
        return f"={quantity}*E{base}"
    return f"={quantity}"


def _write_g(ws: Worksheet, row: int, price: Amount):
    if price is NO_CHARGE:
        _put(ws, f"G{row}", "N/C", fmt=FMT_DETAIL_NUM, font=FONT_DATA)
    elif is_priced(price):
        _put(ws, f"G{row}", f"=E{row}*F{row}", fmt=FMT_DETAIL_NUM,
             font=FONT_DATA)


# --- 공통 꼬리말 --------------------------------------------------------------

def _write_footer(ws: Worksheet, row: int, subtotal_rows: list[int], fmt: str,
                  merges: list[str], lay: Layout, *,
                  always_sum: bool = False):
    """총합계 / 공급가 / 하단 비고 병합 + 인쇄 영역.

    합계가 하나뿐일 때 TOTAL 시트는 `=SUM(G13)` 을, 상세 시트는 `=G10` 을 쓴다.
    """
    _put(ws, f"B{row}", LBL_GRAND, align=CENTER, font=FONT_LABEL)
    _merge_across(merges, "B", "F", row)
    if len(subtotal_rows) == 1 and not always_sum:
        g_formula = f"=G{subtotal_rows[0]}"
    else:
        g_formula = f"=SUM({','.join(f'G{r}' for r in subtotal_rows)})"
    _put(ws, f"G{row}", g_formula, fmt=fmt, font=FONT_DATA_BOLD)
    lay.grand_row = row
    lay.bf_rows.append(row)
    row += 1

    # 공급가는 언제나 공란이다 (골든 전부 그러하다). 수기로 적는 칸이다.
    _put(ws, f"B{row}", LBL_SUPPLY, align=CENTER, font=FONT_LABEL)
    _merge_across(merges, "B", "F", row)
    lay.supply_row = row
    lay.bf_rows.append(row)
    row += 1

    merges.append(f"B{row}:H{row + TRAILER_ROWS - 1}")
    ws.print_area = f"$A$1:$H${row + TRAILER_ROWS - 1}"


# --- 진입점 ------------------------------------------------------------------

def _finish(ws: Worksheet, merges: list[str], lay: Layout) -> None:
    """장식 후 병합. 순서를 지켜야 골든과 같아진다."""
    decorate(ws, lay)
    for ref in merges:
        ws.merge_cells(ref)


def build(quote: Quotation, template: str | Path | IO[bytes],
          *, today: dt.date | None = None) -> Workbook:
    """템플릿(경로 또는 열린 바이트 스트림) 위에 견적서를 그려 워크북을 만든다."""
    today = today or dt.date.today()
    wb = load_workbook(template)

    total_ws = wb[SHEET_TOTAL]
    template_ws = wb[SHEET_TEMPLATE]

    # 템플릿의 H7('DESCRIPTION') 을 지우고 H6:H7 을 병합한다. 복사 전에 해야
    # 상세 시트와 잔존 template 시트가 모두 골든과 같아진다.
    template_ws["H7"].value = None
    template_ws.merge_cells("H6:H7")

    _finish(total_ws, *_write_total_sheet(total_ws, quote, today))

    details = []
    for group in quote.groups:
        ws = wb.copy_worksheet(template_ws)
        ws.title = group.sheet_name
        _finish(ws, *_write_detail_sheet(ws, group, today))
        details.append(ws)

    template_ws.sheet_state = "hidden"
    wb._sheets = [total_ws, *details, template_ws]
    return wb


def build_bytes(quote: Quotation, template_bytes: bytes,
                *, today: dt.date | None = None) -> bytes:
    """템플릿 바이트 -> 완성된 견적서 .xlsx 바이트. 파일시스템을 쓰지 않는다.

    Worker 가 쓰는 진입점이다. 도형 이식까지 마친 최종 바이트를 돌려준다.
    """
    wb = build(quote, BytesIO(template_bytes), today=today)
    buffer = BytesIO()
    wb.save(buffer)
    saved = buffer.getvalue()
    # openpyxl 은 저장 시 그림·도형을 버린다. TOTAL 시트 상단의 로고와
    # 머리글 도형을 템플릿에서 다시 옮겨 붙인다 (SPEC_CELLMAP.md §5.1).
    return drawings.carry_over_bytes(template_bytes, saved, SHEET_TOTAL) or saved


def write(quote: Quotation, template: str | Path, out_path: str | Path,
          *, today: dt.date | None = None) -> Path:
    """`build_bytes` 의 파일 경로 어댑터 (데스크톱)."""
    wb = build(quote, template, today=today)
    out = Path(out_path)
    wb.save(out)
    drawings.carry_over(template, out, SHEET_TOTAL)
    return out
